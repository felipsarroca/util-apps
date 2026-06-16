from __future__ import annotations

import json
import re
import unicodedata
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
SOURCE = next(path for path in ROOT.glob("*.xlsx") if not path.name.startswith("~$"))
OUTPUT = ROOT / "app" / "src" / "SeedData.js"

COURSES = {
    "I3": ("I3", "Infantil", 1, 30),
    "I4": ("I4", "Infantil", 2, 30),
    "I5": ("I5", "Infantil", 3, 30),
    "1CI": ("1r de primària", "Primària", 4, 30),
    "2CI": ("2n de primària", "Primària", 5, 30),
    "3CM": ("3r de primària", "Primària", 6, 30),
    "4CM": ("4t de primària", "Primària", 7, 30),
    "5CS": ("5è de primària", "Primària", 8, 30),
    "6CS": ("6è de primària", "Primària", 9, 30),
    "1ESO": ("1r d'ESO", "Secundària", 10, 33),
    "2ESO": ("2n d'ESO", "Secundària", 11, 33),
    "3ESO": ("3r d'ESO", "Secundària", 12, 33),
    "4ESO": ("4t d'ESO", "Secundària", 13, 33),
}

CHARGE_COLUMNS = {
    "W": "Direcció",
    "X": "Equip directiu",
    "Y": "Coordinació",
    "Z": "DOIP/SIEI",
    "AA": "Patis ESO",
    "AB": "Socialització",
    "AC": "Pastoral",
    "AD": "Teatre",
    "AE": "TIC",
    "AF": "Comitè d'empresa",
    "AG": "Comunicació",
}

TYPE_MAP = {
    "C": ("CLASSE", 1),
    "R": ("REFORC", 0),
    "D": ("DESDOBLAMENT", 0.5),
    "Com": ("COMPLEMENTARIA", 1),
    None: ("CLASSE", 1),
}

EXCLUDED_HEADERS = {"HDC Alumnes", "HDC Famílies", "Total lectives", "Total HDC"}


def normalize(value: object) -> str:
    text = str(value or "").strip().lower()
    text = "".join(
        character
        for character in unicodedata.normalize("NFD", text)
        if unicodedata.category(character) != "Mn"
    )
    return re.sub(r"\s+", " ", text)


def slug(value: object) -> str:
    text = normalize(value)
    text = re.sub(r"[^a-z0-9]+", "-", text).strip("-")
    return text or "registre"


def teacher_name(sheet, row: int) -> str:
    parts = [sheet.cell(row, column).value for column in (3, 1, 2)]
    return " ".join(str(part).strip() for part in parts if part not in (None, "")).strip()


def split_teacher(sheet, row: int) -> dict:
    return {
        "id": f"prof-{row:02d}",
        "cognom1": str(sheet.cell(row, 1).value or "").strip(),
        "cognom2": str(sheet.cell(row, 2).value or "").strip(),
        "nom": str(sheet.cell(row, 3).value or "").strip(),
        "nomComplet": teacher_name(sheet, row),
        "actiu": True,
    }


def subject_target(course_code: str, name: str, raw_target: object) -> float:
    if name == "Tutoria":
        if course_code in {"I3", "I4", "I5"}:
            return 0
        return 1
    if isinstance(raw_target, (int, float)):
        return raw_target
    return 0


def main() -> None:
    formulas = openpyxl.load_workbook(SOURCE, data_only=False)
    values = openpyxl.load_workbook(SOURCE, data_only=True)

    master = values["I3"]
    teachers = [split_teacher(master, row) for row in range(5, 40) if teacher_name(master, row)]
    teacher_by_name = {normalize(item["nomComplet"]): item["id"] for item in teachers}

    courses = []
    subjects_by_key: dict[str, dict] = {}
    plans = []
    assignments = []

    for code, (display_name, stage, order, target_hours) in COURSES.items():
        course_id = f"grup-{slug(code)}"
        courses.append(
            {
                "id": course_id,
                "codi": code,
                "nom": display_name,
                "etapa": stage,
                "ordre": order,
                "horesObjectiu": target_hours,
                "actiu": True,
            }
        )

        formula_sheet = formulas[code]
        value_sheet = values[code]
        current_subject = None
        current_subject_id = None
        plan_seen: set[str] = set()

        for column in range(4, formula_sheet.max_column + 1):
            header = formula_sheet.cell(3, column).value
            if header is not None:
                current_subject = str(header).strip()
                current_subject_id = f"mat-{slug(current_subject)}"
            if not current_subject or current_subject in EXCLUDED_HEADERS:
                continue

            # La clau basada en l'slug unifica variants tipogràfiques com
            # "Ed. Emocional" / "Ed.Emocional" sense duplicar la matèria.
            subject_key = slug(current_subject)
            subjects_by_key.setdefault(
                subject_key,
                {
                    "id": current_subject_id,
                    "nom": current_subject,
                    "nomCurt": current_subject[:28],
                    "activa": True,
                },
            )

            if current_subject_id not in plan_seen:
                raw_target = value_sheet.cell(42, column).value
                plans.append(
                    {
                        "id": f"pla-{slug(code)}-{slug(current_subject)}",
                        "grupId": course_id,
                        "materiaId": current_subject_id,
                        "ordre": len(plan_seen) + 1,
                        "horesObjectiu": subject_target(code, current_subject, raw_target),
                        "tipusBase": TYPE_MAP.get(formula_sheet.cell(4, column).value, ("CLASSE", 1))[0],
                        "actiu": True,
                    }
                )
                plan_seen.add(current_subject_id)

            assignment_type, coverage_factor = TYPE_MAP.get(
                formula_sheet.cell(4, column).value, ("CLASSE", 1)
            )
            for row in range(5, 40):
                hours = value_sheet.cell(row, column).value
                if not isinstance(hours, (int, float)) or hours == 0:
                    continue
                source_name = teacher_name(value_sheet, row)
                teacher_id = teacher_by_name.get(normalize(source_name))
                if not teacher_id:
                    raise RuntimeError(f"Professor no identificat: {source_name} ({code}!{row},{column})")
                assignments.append(
                    {
                        "id": f"ass-{len(assignments) + 1:04d}",
                        "grupId": course_id,
                        "materiaId": current_subject_id,
                        "professorId": teacher_id,
                        "tipus": assignment_type,
                        "hores": hours,
                        "factorCobertura": coverage_factor,
                        "observacions": "",
                        "activa": True,
                    }
                )

    summary = values["Professorat"]
    contracts = []
    charges = [{"id": f"car-{slug(name)}", "nom": name, "ordre": index + 1, "actiu": True}
               for index, name in enumerate(CHARGE_COLUMNS.values())]
    charge_assignments = []

    for row in range(5, 40):
        name = teacher_name(summary, row)
        teacher_id = teacher_by_name.get(normalize(name))
        if not teacher_id:
            continue
        contract_hours = summary.cell(row, 37).value
        if isinstance(contract_hours, (int, float)) and contract_hours > 0:
            contracts.append(
                {
                    "id": f"con-{teacher_id}",
                    "professorId": teacher_id,
                    "hores": contract_hours,
                    "actiu": True,
                }
            )
        for column_letter, charge_name in CHARGE_COLUMNS.items():
            hours = summary[f"{column_letter}{row}"].value
            if isinstance(hours, (int, float)) and hours > 0:
                charge_assignments.append(
                    {
                        "id": f"ac-{len(charge_assignments) + 1:03d}",
                        "professorId": teacher_id,
                        "carrecId": f"car-{slug(charge_name)}",
                        "hores": hours,
                        "observacions": "",
                        "activa": True,
                    }
                )

    payload = {
        "academicYear": {"id": "curs-2025-26", "nom": "2025-26", "estat": "ESBORRANY", "actiu": True},
        "teachers": teachers,
        "contracts": contracts,
        "courses": courses,
        "subjects": list(subjects_by_key.values()),
        "plans": plans,
        "assignments": assignments,
        "charges": charges,
        "chargeAssignments": charge_assignments,
        "rules": [
            {"id": "reg-infantil", "etapa": "Infantil", "percentatgeHdc": 0.20, "actiu": True},
            {"id": "reg-primaria", "etapa": "Primària", "percentatgeHdc": 0.20, "actiu": True},
            {"id": "reg-secundaria", "etapa": "Secundària", "percentatgeHdc": 0.25, "actiu": True},
        ],
    }

    content = (
        "/* Fitxer generat automàticament des de l'Excel. No l'editeu manualment. */\n"
        "const SEED_DATA = "
        + json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
        + ";\n"
    )
    OUTPUT.write_text(content, encoding="utf-8")
    print(
        f"Generat {OUTPUT}: {len(teachers)} professors, {len(courses)} cursos, "
        f"{len(subjects_by_key)} matèries i {len(assignments)} assignacions."
    )


if __name__ == "__main__":
    main()
